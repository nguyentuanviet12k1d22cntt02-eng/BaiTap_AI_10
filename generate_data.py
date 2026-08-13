import os
import csv
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Đường dẫn thư mục chuẩn
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
SCRIPTS_DIR = os.path.join(BASE_DIR, "scripts")
DOCS_DIR = os.path.join(BASE_DIR, "docs")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(SCRIPTS_DIR, exist_ok=True)
os.makedirs(DOCS_DIR, exist_ok=True)

# -------------------------------------------------------------
# 1. BÀI TẬP 1: Dữ liệu Doanh thu các chi nhánh
# -------------------------------------------------------------
headers_bt1 = ["Mã CN", "Tên Chi Nhánh", "Khu Vực", "Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ Nhật"]
branches = [
    ("CN001", "Chi nhánh Ba Đình", "Hà Nội", 12500000, 14200000, 13800000, 15600000, 18900000, 24500000, 26800000),
    ("CN002", "Chi nhánh Cầu Giấy", "Hà Nội", 18200000, 19500000, 17800000, 21000000, 25600000, 31200000, 34500000),
    ("CN003", "Chi nhánh Đống Đa", "Hà Nội", 15400000, 16200000, 14900000, 17800000, 20500000, 27800000, 29600000),
    ("CN004", "Chi nhánh Quận 1", "TP.HCM", 25600000, 27800000, 24900000, 29500000, 35600000, 45200000, 49800000),
    ("CN005", "Chi nhánh Quận 3", "TP.HCM", 20100000, 22400000, 21500000, 24800000, 29800000, 38500000, 41200000),
    ("CN006", "Chi nhánh Bình Thạnh", "TP.HCM", 16800000, 18200000, 17500000, 19800000, 23500000, 30500000, 33200000),
    ("CN007", "Chi nhánh Hải Châu", "Đà Nẵng", 11200000, 12800000, 12100000, 13900000, 16800000, 21500000, 23900000),
    ("CN008", "Chi nhánh Thanh Khê", "Đà Nẵng", 9800000, 10500000, 10200000, 11800000, 14200000, 18600000, 20100000),
    ("CN009", "Chi nhánh Ninh Kiều", "Cần Thơ", 10500000, 11800000, 11200000, 12900000, 15400000, 19800000, 21500000),
    ("CN010", "Chi nhánh Biên Hòa", "Đồng Nai", 13400000, 14800000, 14100000, 16200000, 19500000, 25200000, 27600000),
]

with open(os.path.join(DATA_DIR, "bai_tap_1_doanh_thu_sparkline.csv"), "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers_bt1)
    writer.writerows(branches)

# -------------------------------------------------------------
# 2. BÀI TẬP 2: Dữ liệu Đơn hàng đa sản phẩm (Line items sát thực tế)
# -------------------------------------------------------------
headers_bt2 = ["Mã Đơn", "Ngày Đặt", "Tên Khách Hàng", "Số Điện Thoại", "Địa Chỉ Giao Hàng", "Tên Sản Phẩm", "ĐVT", "Số Lượng", "Đơn Giá", "Thành Tiền", "Trạng Thái", "Link File PDF"]

# Dữ liệu đầy đủ (dùng cho logic nhóm và Excel Merge Cells)
orders_structured = [
    {
        "maDon": "DH-2026-001", "ngayDat": "10/08/2026", "tenKH": "Nguyễn Văn An", "sdt": "0988123456",
        "diaChi": "12 Hoàng Hoa Thám, Ba Đình, Hà Nội", "trangThai": "Chờ xuất", "linkPDF": "",
        "items": [
            ["Laptop Dell XPS 15 9530", "Chiếc", 1, 32000000, 32000000],
            ["Chuột không dây Logitech MX Master 3S", "Chiếc", 1, 2100000, 2100000],
            ["Balo chống sốc Targus 15.6 inch", "Chiếc", 1, 850000, 850000]
        ]
    },
    {
        "maDon": "DH-2026-002", "ngayDat": "10/08/2026", "tenKH": "Trần Thị Bích", "sdt": "0903987654",
        "diaChi": "45 Lê Duẩn, Quận 1, TP.HCM", "trangThai": "Chờ xuất", "linkPDF": "",
        "items": [
            ["Màn hình Dell UltraSharp U2723QE 4K", "Chiếc", 2, 8500000, 17000000],
            ["Giá treo màn hình Human Motion T6 Pro", "Bộ", 2, 890000, 1780000],
            ["Cáp HDMI 2.1 8K Baseus 2m", "Sợi", 2, 250000, 500000]
        ]
    },
    {
        "maDon": "DH-2026-003", "ngayDat": "11/08/2026", "tenKH": "Lê Hoàng Long", "sdt": "0912345678",
        "diaChi": "78 Nguyễn Huệ, Hải Châu, Đà Nẵng", "trangThai": "Chờ xuất", "linkPDF": "",
        "items": [
            ["Bàn phím cơ Keychron K8 Pro RGB", "Chiếc", 2, 2300000, 4600000],
            ["Kê tay bàn phím gỗ óc chó Walnut", "Chiếc", 2, 350000, 700000]
        ]
    },
    {
        "maDon": "DH-2026-004", "ngayDat": "11/08/2026", "tenKH": "Phạm Minh Trang", "sdt": "0977654321",
        "diaChi": "102 Cách Mạng Tháng 8, Quận 3, TP.HCM", "trangThai": "Chờ xuất", "linkPDF": "",
        "items": [
            ["Tai nghe chống ồn Sony WH-1000XM5", "Chiếc", 1, 6800000, 6800000],
            ["Hộp đựng tai nghe cao cấp chống sốc", "Chiếc", 1, 350000, 350000],
            ["Cáp âm thanh Aux 3.5mm Ugreen", "Sợi", 1, 120000, 120000]
        ]
    },
    {
        "maDon": "DH-2026-005", "ngayDat": "11/08/2026", "tenKH": "Đỗ Quang Hưng", "sdt": "0934567890",
        "diaChi": "56 Cầu Giấy, Hà Nội", "trangThai": "Chờ xuất", "linkPDF": "",
        "items": [
            ["Ghế công thái học Ergonomic Sihoo M57", "Chiếc", 1, 4500000, 4500000],
            ["Đệm kê chân công thái học điều chỉnh góc", "Chiếc", 1, 450000, 450000]
        ]
    },
    {
        "maDon": "DH-2026-006", "ngayDat": "09/08/2026", "tenKH": "Vũ Thị Ngọc Hà", "sdt": "0987112233",
        "diaChi": "22 Lý Thường Kiệt, Hoàn Kiếm, Hà Nội", "trangThai": "Đã xuất", "linkPDF": "https://drive.google.com/file/d/demo_phieu_giao_hang/view",
        "items": [
            ["Webcam Elgato Facecam 1080p60", "Chiếc", 1, 3600000, 3600000],
            ["Đèn livestream Elgato Key Light Air", "Chiếc", 1, 2900000, 2900000]
        ]
    },
    {
        "maDon": "DH-2026-007", "ngayDat": "11/08/2026", "tenKH": "Ngô Thành Nam", "sdt": "0908889900",
        "diaChi": "89 Phan Xích Long, Phú Nhuận, TP.HCM", "trangThai": "Chờ xuất", "linkPDF": "",
        "items": [
            ["Ổ cứng di động SSD Samsung T7 Shield 1TB", "Chiếc", 2, 2700000, 5400000],
            ["Hub chuyển đổi USB-C 8 in 1 HyperDrive", "Chiếc", 1, 1850000, 1850000]
        ]
    }
]

# Tạo dữ liệu CSV: Dòng đầu của mỗi đơn ghi đầy đủ thông tin, các dòng phụ bên dưới để trống các cột chung (chỉ 1 ô duy nhất)
orders_csv_rows = []
orders_flat_for_excel = []

for order in orders_structured:
    for idx, it in enumerate(order["items"]):
        if idx == 0:
            # Dòng đầu tiên của đơn: Có đầy đủ thông tin chung
            row_csv = [order["maDon"], order["ngayDat"], order["tenKH"], order["sdt"], order["diaChi"], it[0], it[1], it[2], it[3], it[4], order["trangThai"], order["linkPDF"]]
        else:
            # Các dòng phụ tiếp theo: Để trống thông tin chung để chỉ 1 ô duy nhất hiển thị
            row_csv = ["", "", "", "", "", it[0], it[1], it[2], it[3], it[4], "", ""]
        orders_csv_rows.append(row_csv)
        
        # Dòng cho excel (trước khi merge)
        orders_flat_for_excel.append([order["maDon"], order["ngayDat"], order["tenKH"], order["sdt"], order["diaChi"], it[0], it[1], it[2], it[3], it[4], order["trangThai"], order["linkPDF"]])

with open(os.path.join(DATA_DIR, "bai_tap_2_xuat_hoa_don_pdf.csv"), "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers_bt2)
    writer.writerows(orders_csv_rows)

# -------------------------------------------------------------
# 3. BÀI TẬP 3: Dữ liệu Bảng lương gửi qua Gmail
# -------------------------------------------------------------
headers_bt3 = ["Mã NV", "Họ và Tên", "Phòng Ban", "Email Nhận", "Lương Cơ Bản", "Phụ Cấp", "Thưởng Hiệu Quả", "Khấu Trừ", "Thực Lĩnh", "Trạng Thái", "Thời Gian Gửi"]
payroll = [
    ["NV001", "Hoàng Văn Dũng", "Kinh Doanh", "dung.hoang.demo@gmail.com", 15000000, 2000000, 3500000, 500000, 20000000, "Chưa gửi", ""],
    ["NV002", "Lê Thị Mai", "Kế Toán", "mai.le.demo@gmail.com", 16000000, 1500000, 2000000, 0, 19500000, "Chưa gửi", ""],
    ["NV003", "Trần Quốc Toản", "Kỹ Thuật", "toan.tran.demo@gmail.com", 22000000, 2500000, 4000000, 1000000, 27500000, "Chưa gửi", ""],
    ["NV004", "Nguyễn Hồng Hạnh", "Nhân Sự", "hanh.nguyen.demo@gmail.com", 14000000, 1500000, 1500000, 0, 17000000, "Chưa gửi", ""],
    ["NV005", "Bùi Thanh Tùng", "Marketing", "tung.bui.demo@gmail.com", 17500000, 2000000, 3000000, 500000, 22000000, "Chưa gửi", ""],
    ["NV006", "Đặng Thùy Dương", "Thiết Kế", "duong.dang.demo@gmail.com", 15500000, 1800000, 2200000, 0, 19500000, "Chưa gửi", ""],
    ["NV007", "Phan Anh Tuấn", "Kinh Doanh", "tuan.phan.demo@gmail.com", 13500000, 2000000, 5000000, 800000, 19700000, "Chưa gửi", ""],
]

with open(os.path.join(DATA_DIR, "bai_tap_3_gui_phieu_luong.csv"), "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers_bt3)
    writer.writerows(payroll)

# -------------------------------------------------------------
# 4. BÀI TẬP 4: Dữ liệu Đơn nghỉ phép từ Form
# -------------------------------------------------------------
headers_bt4 = ["Dấu Thời Gian", "Email Nhân Viên", "Họ Tên Nhân Viên", "Phòng Ban", "Số Ngày Nghỉ", "Từ Ngày", "Đến Ngày", "Lý Do Nghỉ", "Mã Đơn", "Trạng Thái", "Quản Lý Duyệt"]
leaves = [
    ["10/08/2026 08:30:15", "dung.hoang.demo@gmail.com", "Hoàng Văn Dũng", "Kinh Doanh", 2, "12/08/2026", "13/08/2026", "Giải quyết việc cá nhân gia đình", "NP-2026-0001", "Chờ Quản Lý Duyệt", ""],
    ["10/08/2026 09:15:22", "mai.le.demo@gmail.com", "Lê Thị Mai", "Kế Toán", 1, "15/08/2026", "15/08/2026", "Khám sức khỏe định kỳ", "NP-2026-0002", "Đã Duyệt", "Trần Trưởng Phòng"],
    ["10/08/2026 10:45:00", "toan.tran.demo@gmail.com", "Trần Quốc Toản", "Kỹ Thuật", 3, "18/08/2026", "20/08/2026", "Nghỉ phép năm về quê", "NP-2026-0003", "Chờ Quản Lý Duyệt", ""],
]

with open(os.path.join(DATA_DIR, "bai_tap_4_don_nghi_phep_form.csv"), "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers_bt4)
    writer.writerows(leaves)

# -------------------------------------------------------------
# 5. BÀI TẬP 5: Dữ liệu Raw 1,000 dòng hỗn tạp cần làm sạch
# -------------------------------------------------------------
headers_bt5 = ["Mã Giao Dịch", "Tên Khách Hàng", "Số Điện Thoại", "Kênh Bán", "Doanh Thu", "Ngày Tạo"]
first_names = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Huỳnh", "Phan", "Vũ", "Võ", "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương"]
mid_names = ["Văn", "Thị", "Hồng", "Minh", "Đức", "Thành", "Ngọc", "Thanh", "Quang", "Anh", "Xuân", "Hoài"]
last_names = ["An", "Bình", "Cường", "Dũng", "Em", "Hương", "Giang", "Hải", "Khánh", "Linh", "Nam", "Phúc", "Quân", "Sơn", "Trang", "Tâm", "Uyên", "Vinh"]
channels = ["Shopee", "Lazada", "Tiki", "TikTok Shop", "Website", "Cửa Hàng Trực Tiếp", "Facebook Ads"]

raw_rows = []
used_codes = []

for i in range(1, 1001):
    if i % 40 == 0:
        ma_gd = ""
    elif i % 25 == 0 and len(used_codes) > 0:
        ma_gd = random.choice(used_codes)
    else:
        ma_gd = f"TRX-2026-{i:05d}"
        used_codes.append(ma_gd)

    name = f"{random.choice(first_names)} {random.choice(mid_names)} {random.choice(last_names)}"
    if i % 7 == 0:
        name = "   " + name.lower() + "  "
    elif i % 5 == 0:
        name = name.upper()

    clean_phone = f"09{random.randint(10000000, 99999999)}"
    if i % 6 == 0:
        phone = clean_phone[1:]
    elif i % 8 == 0:
        phone = f"{clean_phone[:4]}.{clean_phone[4:7]}.{clean_phone[7:]}"
    elif i % 9 == 0:
        phone = f"{clean_phone[:4]} {clean_phone[4:7]} {clean_phone[7:]}"
    else:
        phone = clean_phone

    if i % 50 == 0:
        rev = 0
    elif i % 70 == 0:
        rev = -500000
    else:
        rev = random.randint(150, 8500) * 1000

    channel = random.choice(channels)
    date_val = (datetime(2026, 8, 1) + timedelta(days=random.randint(0, 9))).strftime("%d/%m/%Y")
    raw_rows.append([ma_gd, name, phone, channel, rev, date_val])

with open(os.path.join(DATA_DIR, "bai_tap_5_raw_data_1000_rows.csv"), "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers_bt5)
    writer.writerows(raw_rows)

# -------------------------------------------------------------
# 6. BÀI TẬP 6: Dữ liệu lịch sử giao dịch phân tích RFM
# -------------------------------------------------------------
headers_bt6 = ["Mã Đơn Hàng", "Mã Khách Hàng", "Tên Khách Hàng", "Ngày Mua Hàng", "Doanh Thu Đơn"]
customers_list = [
    ("KH001", "Nguyễn Văn An"),
    ("KH002", "Trần Thị Bích"),
    ("KH003", "Lê Hoàng Long"),
    ("KH004", "Phạm Minh Trang"),
    ("KH005", "Đỗ Quang Hưng"),
    ("KH006", "Vũ Thị Ngọc Hà"),
    ("KH007", "Ngô Thành Nam"),
    ("KH008", "Hoàng Văn Dũng"),
    ("KH009", "Lê Thị Mai"),
    ("KH010", "Trần Quốc Toản"),
    ("KH011", "Nguyễn Hồng Hạnh"),
    ("KH012", "Bùi Thanh Tùng"),
    ("KH013", "Đặng Thùy Dương"),
    ("KH014", "Phan Anh Tuấn"),
    ("KH015", "Võ Hoàng Yến"),
    ("KH016", "Đỗ Mỹ Linh"),
    ("KH017", "Nguyễn Cao Kỳ"),
    ("KH018", "Trần Đức Anh"),
    ("KH019", "Lê Minh Triết"),
    ("KH020", "Phạm Thùy Chi"),
]

random.seed(42)
rfm_rows = []
order_id_counter = 1

for ma_kh, ten_kh in customers_list:
    if ma_kh in ["KH001", "KH002", "KH003"]:
        num_orders = random.randint(10, 15)
    elif ma_kh in ["KH004", "KH005", "KH006", "KH007"]:
        num_orders = random.randint(4, 7)
    elif ma_kh in ["KH008", "KH009", "KH010", "KH011"]:
        num_orders = random.randint(2, 3)
    else:
        num_orders = 1
        
    for _ in range(num_orders):
        ma_don = f"DH-RFM-{order_id_counter:04d}"
        order_id_counter += 1
        
        if ma_kh in ["KH001", "KH002", "KH003", "KH004"]:
            days_ago = random.randint(1, 45)
        elif ma_kh in ["KH005", "KH006", "KH007", "KH008"]:
            days_ago = random.randint(46, 120)
        else:
            days_ago = random.randint(121, 240)
            
        purchase_date = (datetime(2026, 8, 31) - timedelta(days=days_ago)).strftime("%d/%m/%Y")
        
        if ma_kh in ["KH001", "KH004", "KH012"]:
            doanh_thu = random.randint(3000, 12000) * 1000
        else:
            doanh_thu = random.randint(300, 4500) * 1000
            
        rfm_rows.append([ma_don, ma_kh, ten_kh, purchase_date, doanh_thu])

rfm_rows.sort(key=lambda x: x[0])

with open(os.path.join(DATA_DIR, "bai_tap_6_rfm_analysis.csv"), "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers_bt6)
    writer.writerows(rfm_rows)

# -------------------------------------------------------------
# 7. BÀI TẬP 7: Hệ thống Quản lý Shopping HOÀN CHỈNH - DỮ LIỆU MỞ RỘNG
# -------------------------------------------------------------

# 7.1. Sheet Danh Mục Sản Phẩm (Categories)
headers_bt7_categories = ["Mã Danh Mục", "Tên Danh Mục", "Mô Tả"]
categories_rows = [
    ["DM001", "Laptop & Máy Tính", "Laptop, PC, Workstation các loại"],
    ["DM002", "Màn Hình & Thiết Bị Hiển Thị", "Màn hình máy tính, TV, Projector"],
    ["DM003", "Bàn Phím & Chuột", "Bàn phím cơ, chuột gaming, chuột văn phòng"],
    ["DM004", "Âm Thanh", "Tai nghe, loa, micro, sound card"],
    ["DM005", "Nội Thất Văn Phòng", "Ghế, bàn, giá đỡ, phụ kiện văn phòng"],
    ["DM006", "Thiết Bị Quay Phim & Streaming", "Webcam, micro thu âm, đèn led"],
    ["DM007", "Phụ Kiện & Linh Kiện", "Cáp, adapter, ổ cứng, RAM, SSD"]
]

# 7.2. Sheet Sản Phẩm (Products) - Mở rộng lên 54 sản phẩm (x3)
headers_bt7_products = ["Mã Sản Phẩm", "Tên Sản Phẩm", "Mã Danh Mục", "ĐVT", "Đơn Giá (VNĐ)", "Tồn Kho", "Tồn Kho Tối Thiểu"]
products_rows = [
    # Laptop & Máy Tính (15 sản phẩm)
    ["SP001", "Laptop Dell XPS 15 9530", "DM001", "Chiếc", 32000000, 15, 5],
    ["SP002", "Laptop HP Envy 13", "DM001", "Chiếc", 24000000, 8, 3],
    ["SP003", "Laptop Asus ROG Zephyrus G14", "DM001", "Chiếc", 28000000, 12, 5],
    ["SP004", "Laptop MacBook Air M2", "DM001", "Chiếc", 27000000, 18, 8],
    ["SP005", "Laptop Lenovo ThinkPad X1 Carbon", "DM001", "Chiếc", 35000000, 10, 4],
    ["SP006", "Laptop Acer Swift 3", "DM001", "Chiếc", 18000000, 20, 8],
    ["SP007", "Laptop MSI GF63 Thin", "DM001", "Chiếc", 19500000, 14, 6],
    ["SP008", "PC Gaming Custom RTX 4070", "DM001", "Bộ", 45000000, 5, 2],
    ["SP009", "Mac Mini M2 Pro", "DM001", "Chiếc", 22000000, 9, 3],
    ["SP010", "Laptop Dell Inspiron 15", "DM001", "Chiếc", 16000000, 25, 10],
    ["SP011", "Laptop HP Pavilion Gaming", "DM001", "Chiếc", 21000000, 11, 5],
    ["SP012", "Laptop Lenovo Legion 5", "DM001", "Chiếc", 26000000, 13, 6],
    ["SP013", "Laptop Asus Vivobook 14", "DM001", "Chiếc", 14000000, 22, 10],
    ["SP014", "Laptop Microsoft Surface Laptop 5", "DM001", "Chiếc", 29000000, 7, 3],
    ["SP015", "Laptop Gigabyte Aorus 15", "DM001", "Chiếc", 33000000, 6, 2],
    
    # Màn Hình (9 sản phẩm)
    ["SP016", "Màn hình Dell UltraSharp 27 inch 4K", "DM002", "Chiếc", 8500000, 24, 10],
    ["SP017", "Màn hình LG 24 inch Full HD", "DM002", "Chiếc", 3200000, 35, 15],
    ["SP018", "Màn hình Samsung Odyssey G7 32 inch", "DM002", "Chiếc", 12000000, 16, 8],
    ["SP019", "Màn hình Asus ProArt PA279CV 27 inch", "DM002", "Chiếc", 9500000, 18, 8],
    ["SP020", "Màn hình AOC 24G2 Gaming", "DM002", "Chiếc", 4200000, 28, 12],
    ["SP021", "Màn hình BenQ PD2700U 27 inch 4K", "DM002", "Chiếc", 11000000, 14, 6],
    ["SP022", "Màn hình ViewSonic VX2418 24 inch", "DM002", "Chiếc", 3800000, 30, 15],
    ["SP023", "Màn hình LG UltraWide 34 inch", "DM002", "Chiếc", 15000000, 10, 5],
    ["SP024", "Màn hình Xiaomi 27 inch 2K", "DM002", "Chiếc", 4500000, 22, 10],
    
    # Bàn Phím & Chuột (12 sản phẩm)
    ["SP025", "Bàn phím cơ Keychron K8 Pro RGB", "DM003", "Chiếc", 2300000, 45, 20],
    ["SP026", "Bàn phím Logitech K380 Bluetooth", "DM003", "Chiếc", 650000, 60, 25],
    ["SP027", "Bàn phím cơ Royal Kludge RK61", "DM003", "Chiếc", 980000, 52, 22],
    ["SP028", "Bàn phím Corsair K70 RGB", "DM003", "Chiếc", 3200000, 38, 18],
    ["SP029", "Bàn phím Leopold FC660C", "DM003", "Chiếc", 5500000, 15, 8],
    ["SP030", "Bàn phím Akko 3098B", "DM003", "Chiếc", 1250000, 48, 20],
    ["SP031", "Chuột Logitech MX Master 3S", "DM003", "Chiếc", 2100000, 30, 15],
    ["SP032", "Chuột gaming Razer DeathAdder V3", "DM003", "Chiếc", 1500000, 28, 12],
    ["SP033", "Chuột Logitech G502 Hero", "DM003", "Chiếc", 1100000, 42, 18],
    ["SP034", "Chuột Apple Magic Mouse", "DM003", "Chiếc", 1800000, 25, 12],
    ["SP035", "Chuột Corsair Dark Core RGB Pro", "DM003", "Chiếc", 2400000, 20, 10],
    ["SP036", "Chuột gaming SteelSeries Rival 3", "DM003", "Chiếc", 750000, 50, 22],
    
    # Âm Thanh (9 sản phẩm)
    ["SP037", "Tai nghe Sony WH-1000XM5", "DM004", "Chiếc", 6800000, 18, 8],
    ["SP038", "Tai nghe Apple AirPods Pro 2", "DM004", "Chiếc", 5200000, 25, 10],
    ["SP039", "Tai nghe Bose QuietComfort 45", "DM004", "Chiếc", 7200000, 15, 7],
    ["SP040", "Tai nghe gaming HyperX Cloud II", "DM004", "Chiếc", 1800000, 32, 15],
    ["SP041", "Tai nghe Sennheiser HD 560S", "DM004", "Chiếc", 4500000, 12, 6],
    ["SP042", "Loa Bluetooth JBL Flip 6", "DM004", "Chiếc", 2800000, 28, 12],
    ["SP043", "Loa Edifier S880DB", "DM004", "Đôi", 5500000, 10, 5],
    ["SP044", "Micro thu âm Audio-Technica AT2020", "DM004", "Chiếc", 2400000, 18, 8],
    ["SP045", "Soundbar Samsung HW-Q600A", "DM004", "Chiếc", 9500000, 8, 4],
    
    # Nội Thất Văn Phòng (6 sản phẩm)
    ["SP046", "Ghế công thái học Sihoo M57", "DM005", "Chiếc", 4500000, 12, 5],
    ["SP047", "Bàn làm việc điều chỉnh chiều cao", "DM005", "Chiếc", 6800000, 6, 3],
    ["SP048", "Ghế Herman Miller Aeron (Refurbished)", "DM005", "Chiếc", 18000000, 3, 1],
    ["SP049", "Giá treo màn hình Human Motion T6 Pro", "DM005", "Bộ", 890000, 45, 20],
    ["SP050", "Đèn bàn Xiaomi Mi Smart LED", "DM005", "Chiếc", 850000, 38, 18],
    ["SP051", "Kệ sách gỗ 5 tầng", "DM005", "Chiếc", 2200000, 15, 6],
    
    # Thiết Bị Streaming (6 sản phẩm)
    ["SP052", "Webcam Elgato Facecam 1080p60", "DM006", "Chiếc", 3600000, 10, 5],
    ["SP053", "Micro thu âm Blue Yeti X", "DM006", "Chiếc", 4200000, 8, 4],
    ["SP054", "Đèn Elgato Key Light Air", "DM006", "Chiếc", 3800000, 12, 6],
    ["SP055", "Stream Deck Elgato 15 Keys", "DM006", "Chiếc", 3200000, 9, 4],
    ["SP056", "Green Screen Elgato", "DM006", "Chiếc", 4500000, 6, 3],
    ["SP057", "Ring Light 18 inch với chân đỡ", "DM006", "Bộ", 1500000, 20, 10],
    
    # Phụ Kiện & Linh Kiện (9 sản phẩm)
    ["SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "DM007", "Chiếc", 2700000, 40, 20],
    ["SP059", "Hub USB-C 8 in 1 HyperDrive", "DM007", "Chiếc", 1850000, 55, 25],
    ["SP060", "Cáp HDMI 2.1 8K Baseus 2m", "DM007", "Sợi", 250000, 120, 50],
    ["SP061", "Balo chống sốc Targus 15.6 inch", "DM007", "Chiếc", 850000, 32, 15],
    ["SP062", "RAM Corsair Vengeance 32GB DDR4", "DM007", "Bộ", 3200000, 25, 12],
    ["SP063", "SSD Kingston NV2 1TB NVMe", "DM007", "Chiếc", 1800000, 48, 22],
    ["SP064", "Ổ cứng WD My Passport 5TB", "DM007", "Chiếc", 3500000, 22, 10],
    ["SP065", "Adapter Anker USB-C to HDMI", "DM007", "Chiếc", 580000, 65, 30],
    ["SP066", "Đế tản nhiệt laptop Cooler Master", "DM007", "Chiếc", 650000, 42, 20]
]

# 7.3. Sheet Khách Hàng (Customers) - Mở rộng lên 30 khách (x3)
headers_bt7_customers = ["Mã Khách Hàng", "Tên Khách Hàng", "Số Điện Thoại", "Email", "Địa Chỉ", "Thành Phố", "Ngày Đăng Ký", "Loại Khách Hàng"]
customers_rows = [
    ["KH001", "Nguyen Van An", "0988123456", "nguyenvanan@gmail.com", "12 Hoang Hoa Tham, Ba Dinh", "Ha Noi", "15/07/2026", "VIP"],
    ["KH002", "Tran Thi Bich", "0903987654", "tranbich@gmail.com", "45 Le Duan, Quan 1", "TP.HCM", "20/07/2026", "Thường"],
    ["KH003", "Le Hoang Long", "0912345678", "lehoanglong@gmail.com", "78 Nguyen Hue, Hai Chau", "Da Nang", "22/07/2026", "Thường"],
    ["KH004", "Pham Minh Trang", "0977654321", "phamtrang@gmail.com", "102 Cach Mang Thang 8, Quan 3", "TP.HCM", "25/07/2026", "VIP"],
    ["KH005", "Do Quang Hung", "0934567890", "doquanghung@gmail.com", "56 Cau Giay", "Ha Noi", "28/07/2026", "Thường"],
    ["KH006", "Vu Thi Ngoc Ha", "0987112233", "vungocha@gmail.com", "22 Ly Thuong Kiet, Hoan Kiem", "Ha Noi", "01/08/2026", "VIP"],
    ["KH007", "Ngo Thanh Nam", "0908889900", "ngothanhnam@gmail.com", "15 Tran Phu, Ngo Quyen", "Hai Phong", "03/08/2026", "Thường"],
    ["KH008", "Hoang Van Duc", "0965443322", "hoangduc@gmail.com", "120 Nguyen Van Cu, Ninh Kieu", "Can Tho", "05/08/2026", "Thường"],
    ["KH009", "Nguyen Thi Mai", "0943221100", "nguyenmai@gmail.com", "88 Le Loi", "Vinh", "07/08/2026", "Thường"],
    ["KH010", "Bui Anh Tuan", "0921998877", "buianhtuan@gmail.com", "50 Quang Trung, Hai Chau", "Da Nang", "09/08/2026", "VIP"],
    ["KH011", "Tran Minh Chau", "0976554433", "tranminhchau@gmail.com", "234 Tran Hung Dao, Quan 5", "TP.HCM", "12/07/2026", "VIP"],
    ["KH012", "Phan Thi Lan", "0988776655", "phantlan@gmail.com", "67 Bach Dang, Hai Ba Trung", "Ha Noi", "18/07/2026", "Thường"],
    ["KH013", "Dang Hoai Nam", "0912667788", "danghoinam@gmail.com", "45 Le Loi, Thanh Khe", "Da Nang", "21/07/2026", "Thường"],
    ["KH014", "Vo Thi Thu", "0934889922", "vothithu@gmail.com", "89 Nguyen Thi Minh Khai, Quan 3", "TP.HCM", "24/07/2026", "VIP"],
    ["KH015", "Nguyen Quang Huy", "0965112233", "nguyenquanghuy@gmail.com", "123 Giang Vo, Ba Dinh", "Ha Noi", "27/07/2026", "Thường"],
    ["KH016", "Le Thi Huong", "0977443322", "lethihuong@gmail.com", "56 Hai Phong, Le Chan", "Hai Phong", "30/07/2026", "Thường"],
    ["KH017", "Hoang Minh Duc", "0908334455", "hoangminhduc@gmail.com", "78 3/2, Ninh Kieu", "Can Tho", "02/08/2026", "VIP"],
    ["KH018", "Pham Van Tuan", "0943667788", "phamvantuan@gmail.com", "234 Le Hong Phong", "Vinh", "04/08/2026", "Thường"],
    ["KH019", "Ngo Thi Huyen", "0921556677", "ngothihuyen@gmail.com", "45 Nguyen Van Linh, Hai Chau", "Da Nang", "06/08/2026", "Thường"],
    ["KH020", "Tran Quoc Toan", "0987334455", "tranquoctoan@gmail.com", "123 Pasteur, Quan 1", "TP.HCM", "08/08/2026", "VIP"],
    ["KH021", "Vu Van Thanh", "0976889922", "vuvanthanh@gmail.com", "67 Tran Phu, Hoan Kiem", "Ha Noi", "10/07/2026", "Thường"],
    ["KH022", "Doan Thi My", "0988112244", "doanthimy@gmail.com", "89 Bach Dang, Hai Chau", "Da Nang", "13/07/2026", "VIP"],
    ["KH023", "Nguyen Thanh Binh", "0912445566", "nguyenthanhbinh@gmail.com", "56 Dong Khoi, Quan 1", "TP.HCM", "16/07/2026", "Thường"],
    ["KH024", "Le Van Khanh", "0934778899", "levankhanh@gmail.com", "234 Giai Phong, Hai Ba Trung", "Ha Noi", "19/07/2026", "Thường"],
    ["KH025", "Phan Thi Nga", "0965223344", "phanthinga@gmail.com", "78 Quang Trung, Hong Bang", "Hai Phong", "23/07/2026", "VIP"],
    ["KH026", "Hoang Thi Loan", "0977556688", "hoangthiloan@gmail.com", "45 Tran Hung Dao, Ninh Kieu", "Can Tho", "26/07/2026", "Thường"],
    ["KH027", "Tran Van Hai", "0908667799", "tranvanhai@gmail.com", "123 Nguyen Du", "Vinh", "29/07/2026", "Thường"],
    ["KH028", "Nguyen Thi Dung", "0943889911", "nguyenthidung@gmail.com", "67 Phan Chu Trinh, Hai Chau", "Da Nang", "31/07/2026", "VIP"],
    ["KH029", "Vo Van Long", "0921445577", "vovanlong@gmail.com", "89 Le Lai, Quan 1", "TP.HCM", "03/08/2026", "Thường"],
    ["KH030", "Dang Thi Phuong", "0987667788", "dangthiphuong@gmail.com", "56 Hai Ba Trung, Ba Dinh", "Ha Noi", "05/08/2026", "VIP"]
]

# 7.4. Sheet Đơn Hàng (Orders) - Mở rộng lên 30 đơn (x3)
headers_bt7_orders = ["Mã Đơn Hàng", "Mã Khách Hàng", "Ngày Đặt", "Trạng Thái", "Tổng Tiền (VNĐ)", "Phương Thức Thanh Toán", "Ghi Chú"]
orders_master = [
    ["DH-001", "KH001", "01/08/2026", "Hoàn Thành", 34100000, "Chuyển Khoản", "Giao nhanh trong ngày"],
    ["DH-002", "KH002", "02/08/2026", "Đang Giao", 11700000, "COD", ""],
    ["DH-003", "KH003", "03/08/2026", "Hoàn Thành", 29300000, "Chuyển Khoản", ""],
    ["DH-004", "KH004", "04/08/2026", "Đang Xử Lý", 10800000, "Thẻ Tín Dụng", "Khách yêu cầu kiểm tra kỹ"],
    ["DH-005", "KH005", "05/08/2026", "Hoàn Thành", 7000000, "Chuyển Khoản", ""],
    ["DH-006", "KH006", "06/08/2026", "Đang Giao", 10400000, "COD", "Gọi trước khi giao"],
    ["DH-007", "KH007", "07/08/2026", "Hoàn Thành", 7200000, "Chuyển Khoản", ""],
    ["DH-008", "KH008", "08/08/2026", "Hoàn Thành", 4600000, "COD", ""],
    ["DH-009", "KH009", "09/08/2026", "Đang Xử Lý", 8300000, "Chuyển Khoản", ""],
    ["DH-010", "KH010", "10/08/2026", "Đang Giao", 15300000, "Thẻ Tín Dụng", ""],
    ["DH-011", "KH011", "11/08/2026", "Hoàn Thành", 52000000, "Chuyển Khoản", "Khách VIP - Ưu tiên"],
    ["DH-012", "KH012", "11/08/2026", "Hoàn Thành", 3200000, "COD", ""],
    ["DH-013", "KH013", "12/08/2026", "Đang Giao", 18000000, "Chuyển Khoản", ""],
    ["DH-014", "KH014", "12/08/2026", "Hoàn Thành", 12000000, "Thẻ Tín Dụng", ""],
    ["DH-015", "KH015", "13/08/2026", "Đang Xử Lý", 6100000, "Chuyển Khoản", ""],
    ["DH-016", "KH016", "13/08/2026", "Hoàn Thành", 4500000, "COD", ""],
    ["DH-017", "KH017", "14/08/2026", "Đang Giao", 27000000, "Chuyển Khoản", "Giao hàng cẩn thận"],
    ["DH-018", "KH018", "14/08/2026", "Hoàn Thành", 9500000, "COD", ""],
    ["DH-019", "KH019", "15/08/2026", "Đang Xử Lý", 15300000, "Chuyển Khoản", ""],
    ["DH-020", "KH020", "15/08/2026", "Hoàn Thành", 35000000, "Thẻ Tín Dụng", "Khách VIP"],
    ["DH-021", "KH021", "16/08/2026", "Đang Giao", 8200000, "Chuyển Khoản", ""],
    ["DH-022", "KH022", "16/08/2026", "Hoàn Thành", 45000000, "Chuyển Khoản", "Khách VIP - Tặng quà"],
    ["DH-023", "KH023", "17/08/2026", "Đang Xử Lý", 5700000, "COD", ""],
    ["DH-024", "KH024", "17/08/2026", "Hoàn Thành", 19500000, "Chuyển Khoản", ""],
    ["DH-025", "KH025", "18/08/2026", "Đang Giao", 12500000, "Thẻ Tín Dụng", ""],
    ["DH-026", "KH026", "18/08/2026", "Hoàn Thành", 3850000, "COD", ""],
    ["DH-027", "KH027", "19/08/2026", "Đang Xử Lý", 24000000, "Chuyển Khoản", ""],
    ["DH-028", "KH028", "19/08/2026", "Hoàn Thành", 33000000, "Chuyển Khoản", "Khách VIP"],
    ["DH-029", "KH029", "20/08/2026", "Đang Giao", 16800000, "COD", ""],
    ["DH-030", "KH030", "20/08/2026", "Hoàn Thành", 28000000, "Thẻ Tín Dụng", "Khách VIP - Giao express"]
]

# 7.5. Sheet Chi Tiết Đơn Hàng (OrderDetails) - Mở rộng lên ~60 dòng (x4 vì có đơn nhiều sp)
headers_bt7_order_details = ["Mã Chi Tiết", "Mã Đơn Hàng", "Mã Sản Phẩm", "Tên Sản Phẩm", "ĐVT", "Số Lượng", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)"]
order_details_rows = [
    # DH-001 (2 sản phẩm)
    ["CT001", "DH-001", "SP001", "Laptop Dell XPS 15 9530", "Chiếc", 1, 32000000, 32000000],
    ["CT002", "DH-001", "SP031", "Chuột Logitech MX Master 3S", "Chiếc", 1, 2100000, 2100000],
    # DH-002 (2 sản phẩm)
    ["CT003", "DH-002", "SP016", "Màn hình Dell UltraSharp 27 inch 4K", "Chiếc", 1, 8500000, 8500000],
    ["CT004", "DH-002", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Chiếc", 1, 2300000, 2300000],
    ["CT005", "DH-002", "SP060", "Cáp HDMI 2.1 8K Baseus 2m", "Sợi", 4, 250000, 1000000],
    # DH-003 (2 sản phẩm)
    ["CT006", "DH-003", "SP004", "Laptop MacBook Air M2", "Chiếc", 1, 27000000, 27000000],
    ["CT007", "DH-003", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Chiếc", 1, 2300000, 2300000],
    # DH-004 (3 sản phẩm)
    ["CT008", "DH-004", "SP031", "Chuột Logitech MX Master 3S", "Chiếc", 3, 2100000, 6300000],
    ["CT009", "DH-004", "SP046", "Ghế công thái học Sihoo M57", "Chiếc", 1, 4500000, 4500000],
    # DH-005 (2 sản phẩm)
    ["CT010", "DH-005", "SP037", "Tai nghe Sony WH-1000XM5", "Chiếc", 1, 6800000, 6800000],
    ["CT011", "DH-005", "SP060", "Cáp HDMI 2.1 8K Baseus 2m", "Sợi", 1, 250000, 250000],
    # DH-006 (2 sản phẩm)
    ["CT012", "DH-006", "SP037", "Tai nghe Sony WH-1000XM5", "Chiếc", 1, 6800000, 6800000],
    ["CT013", "DH-006", "SP052", "Webcam Elgato Facecam 1080p60", "Chiếc", 1, 3600000, 3600000],
    # DH-007 (2 sản phẩm)
    ["CT014", "DH-007", "SP052", "Webcam Elgato Facecam 1080p60", "Chiếc", 2, 3600000, 7200000],
    # DH-008 (2 sản phẩm)
    ["CT015", "DH-008", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Chiếc", 2, 2300000, 4600000],
    # DH-009 (3 sản phẩm)
    ["CT016", "DH-009", "SP031", "Chuột Logitech MX Master 3S", "Chiếc", 1, 2100000, 2100000],
    ["CT017", "DH-009", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Chiếc", 2, 2700000, 5400000],
    ["CT018", "DH-009", "SP061", "Balo chống sốc Targus 15.6 inch", "Chiếc", 1, 850000, 850000],
    # DH-010 (2 sản phẩm)
    ["CT019", "DH-010", "SP016", "Màn hình Dell UltraSharp 27 inch 4K", "Chiếc", 1, 8500000, 8500000],
    ["CT020", "DH-010", "SP037", "Tai nghe Sony WH-1000XM5", "Chiếc", 1, 6800000, 6800000],
    # DH-011 (3 sản phẩm - Đơn lớn)
    ["CT021", "DH-011", "SP005", "Laptop Lenovo ThinkPad X1 Carbon", "Chiếc", 1, 35000000, 35000000],
    ["CT022", "DH-011", "SP023", "Màn hình LG UltraWide 34 inch", "Chiếc", 1, 15000000, 15000000],
    ["CT023", "DH-011", "SP031", "Chuột Logitech MX Master 3S", "Chiếc", 1, 2100000, 2100000],
    # DH-012 (1 sản phẩm)
    ["CT024", "DH-012", "SP017", "Màn hình LG 24 inch Full HD", "Chiếc", 1, 3200000, 3200000],
    # DH-013 (2 sản phẩm)
    ["CT025", "DH-013", "SP002", "Laptop HP Envy 13", "Chiếc", 1, 18000000, 18000000],
    # DH-014 (2 sản phẩm)
    ["CT026", "DH-014", "SP018", "Màn hình Samsung Odyssey G7 32 inch", "Chiếc", 1, 12000000, 12000000],
    # DH-015 (3 sản phẩm)
    ["CT027", "DH-015", "SP028", "Bàn phím Corsair K70 RGB", "Chiếc", 1, 3200000, 3200000],
    ["CT028", "DH-015", "SP042", "Loa Bluetooth JBL Flip 6", "Chiếc", 1, 2800000, 2800000],
    ["CT029", "DH-015", "SP065", "Adapter Anker USB-C to HDMI", "Chiếc", 2, 580000, 1160000],
    # DH-016 (1 sản phẩm)
    ["CT030", "DH-016", "SP046", "Ghế công thái học Sihoo M57", "Chiếc", 1, 4500000, 4500000],
    # DH-017 (1 sản phẩm)
    ["CT031", "DH-017", "SP004", "Laptop MacBook Air M2", "Chiếc", 1, 27000000, 27000000],
    # DH-018 (2 sản phẩm)
    ["CT032", "DH-018", "SP037", "Tai nghe Sony WH-1000XM5", "Chiếc", 1, 6800000, 6800000],
    ["CT033", "DH-018", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Chiếc", 1, 2700000, 2700000],
    # DH-019 (3 sản phẩm)
    ["CT034", "DH-019", "SP021", "Màn hình BenQ PD2700U 27 inch 4K", "Chiếc", 1, 11000000, 11000000],
    ["CT035", "DH-019", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Chiếc", 1, 2300000, 2300000],
    ["CT036", "DH-019", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Chiếc", 1, 2700000, 2700000],
    # DH-020 (1 sản phẩm - VIP)
    ["CT037", "DH-020", "SP005", "Laptop Lenovo ThinkPad X1 Carbon", "Chiếc", 1, 35000000, 35000000],
    # DH-021 (2 sản phẩm)
    ["CT038", "DH-021", "SP039", "Tai nghe Bose QuietComfort 45", "Chiếc", 1, 7200000, 7200000],
    ["CT039", "DH-021", "SP060", "Cáp HDMI 2.1 8K Baseus 2m", "Sợi", 4, 250000, 1000000],
    # DH-022 (1 sản phẩm - VIP, đơn lớn)
    ["CT040", "DH-022", "SP008", "PC Gaming Custom RTX 4070", "Bộ", 1, 45000000, 45000000],
    # DH-023 (3 sản phẩm)
    ["CT041", "DH-023", "SP026", "Bàn phím Logitech K380 Bluetooth", "Chiếc", 2, 650000, 1300000],
    ["CT042", "DH-023", "SP036", "Chuột gaming SteelSeries Rival 3", "Chiếc", 2, 750000, 1500000],
    ["CT043", "DH-023", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Chiếc", 1, 2700000, 2700000],
    ["CT044", "DH-023", "SP065", "Adapter Anker USB-C to HDMI", "Chiếc", 3, 580000, 1740000],
    # DH-024 (2 sản phẩm)
    ["CT045", "DH-024", "SP007", "Laptop MSI GF63 Thin", "Chiếc", 1, 19500000, 19500000],
    # DH-025 (3 sản phẩm)
    ["CT046", "DH-025", "SP040", "Tai nghe gaming HyperX Cloud II", "Chiếc", 2, 1800000, 3600000],
    ["CT047", "DH-025", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Chiếc", 3, 2700000, 8100000],
    ["CT048", "DH-025", "SP061", "Balo chống sốc Targus 15.6 inch", "Chiếc", 1, 850000, 850000],
    # DH-026 (2 sản phẩm)
    ["CT049", "DH-026", "SP052", "Webcam Elgato Facecam 1080p60", "Chiếc", 1, 3600000, 3600000],
    ["CT050", "DH-026", "SP060", "Cáp HDMI 2.1 8K Baseus 2m", "Sợi", 1, 250000, 250000],
    # DH-027 (1 sản phẩm)
    ["CT051", "DH-027", "SP002", "Laptop HP Envy 13", "Chiếc", 1, 24000000, 24000000],
    # DH-028 (1 sản phẩm - VIP)
    ["CT052", "DH-028", "SP015", "Laptop Gigabyte Aorus 15", "Chiếc", 1, 33000000, 33000000],
    # DH-029 (3 sản phẩm)
    ["CT053", "DH-029", "SP019", "Màn hình Asus ProArt PA279CV 27 inch", "Chiếc", 1, 9500000, 9500000],
    ["CT054", "DH-029", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Chiếc", 1, 2300000, 2300000],
    ["CT055", "DH-029", "SP043", "Loa Edifier S880DB", "Đôi", 1, 5500000, 5500000],
    # DH-030 (1 sản phẩm - VIP)
    ["CT056", "DH-030", "SP003", "Laptop Asus ROG Zephyrus G14", "Chiếc", 1, 28000000, 28000000]
]

# 7.6. Sheet Lịch Sử Tồn Kho (InventoryHistory) - Mở rộng lên ~35 giao dịch
headers_bt7_inventory = ["Mã Giao Dịch", "Ngày", "Mã Sản Phẩm", "Tên Sản Phẩm", "Loại", "Số Lượng", "Tồn Sau GD", "Người Thực Hiện", "Ghi Chú"]
inventory_rows = [
    ["GD001", "25/07/2026", "SP001", "Laptop Dell XPS 15 9530", "Nhập Kho", 20, 20, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD002", "26/07/2026", "SP031", "Chuột Logitech MX Master 3S", "Nhập Kho", 50, 50, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD003", "26/07/2026", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Nhập Kho", 60, 60, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD004", "27/07/2026", "SP016", "Màn hình Dell UltraSharp 27 inch 4K", "Nhập Kho", 30, 30, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD005", "27/07/2026", "SP037", "Tai nghe Sony WH-1000XM5", "Nhập Kho", 25, 25, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD006", "28/07/2026", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Nhập Kho", 60, 60, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD007", "28/07/2026", "SP004", "Laptop MacBook Air M2", "Nhập Kho", 22, 22, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD008", "29/07/2026", "SP046", "Ghế công thái học Sihoo M57", "Nhập Kho", 15, 15, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD009", "30/07/2026", "SP052", "Webcam Elgato Facecam 1080p60", "Nhập Kho", 18, 18, "Admin", "Nhập hàng từ nhà cung cấp"],
    ["GD010", "01/08/2026", "SP001", "Laptop Dell XPS 15 9530", "Xuất Bán", -1, 19, "NV001", "Đơn hàng DH-001"],
    ["GD011", "01/08/2026", "SP031", "Chuột Logitech MX Master 3S", "Xuất Bán", -1, 49, "NV001", "Đơn hàng DH-001"],
    ["GD012", "02/08/2026", "SP016", "Màn hình Dell UltraSharp 27 inch 4K", "Xuất Bán", -1, 29, "NV002", "Đơn hàng DH-002"],
    ["GD013", "02/08/2026", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Xuất Bán", -1, 59, "NV002", "Đơn hàng DH-002"],
    ["GD014", "03/08/2026", "SP004", "Laptop MacBook Air M2", "Xuất Bán", -1, 21, "NV001", "Đơn hàng DH-003"],
    ["GD015", "03/08/2026", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Xuất Bán", -1, 58, "NV001", "Đơn hàng DH-003"],
    ["GD016", "04/08/2026", "SP031", "Chuột Logitech MX Master 3S", "Xuất Bán", -3, 46, "NV003", "Đơn hàng DH-004"],
    ["GD017", "04/08/2026", "SP046", "Ghế công thái học Sihoo M57", "Xuất Bán", -1, 14, "NV003", "Đơn hàng DH-004"],
    ["GD018", "05/08/2026", "SP037", "Tai nghe Sony WH-1000XM5", "Xuất Bán", -1, 24, "NV002", "Đơn hàng DH-005"],
    ["GD019", "06/08/2026", "SP037", "Tai nghe Sony WH-1000XM5", "Xuất Bán", -1, 23, "NV001", "Đơn hàng DH-006"],
    ["GD020", "06/08/2026", "SP052", "Webcam Elgato Facecam 1080p60", "Xuất Bán", -1, 17, "NV001", "Đơn hàng DH-006"],
    ["GD021", "07/08/2026", "SP052", "Webcam Elgato Facecam 1080p60", "Xuất Bán", -2, 15, "NV002", "Đơn hàng DH-007"],
    ["GD022", "08/08/2026", "SP025", "Bàn phím cơ Keychron K8 Pro RGB", "Xuất Bán", -2, 56, "NV003", "Đơn hàng DH-008"],
    ["GD023", "09/08/2026", "SP031", "Chuột Logitech MX Master 3S", "Xuất Bán", -1, 45, "NV001", "Đơn hàng DH-009"],
    ["GD024", "09/08/2026", "SP058", "Ổ cứng SSD Samsung T7 Shield 1TB", "Xuất Bán", -2, 58, "NV001", "Đơn hàng DH-009"],
    ["GD025", "10/08/2026", "SP016", "Màn hình Dell UltraSharp 27 inch 4K", "Xuất Bán", -1, 28, "NV002", "Đơn hàng DH-010"],
    ["GD026", "10/08/2026", "SP037", "Tai nghe Sony WH-1000XM5", "Xuất Bán", -1, 22, "NV002", "Đơn hàng DH-010"],
    ["GD027", "11/08/2026", "SP005", "Laptop Lenovo ThinkPad X1 Carbon", "Nhập Kho", 15, 15, "Admin", "Bổ sung tồn kho"],
    ["GD028", "11/08/2026", "SP023", "Màn hình LG UltraWide 34 inch", "Nhập Kho", 12, 12, "Admin", "Bổ sung tồn kho"],
    ["GD029", "12/08/2026", "SP008", "PC Gaming Custom RTX 4070", "Nhập Kho", 8, 8, "Admin", "Đặt hàng theo yêu cầu VIP"],
    ["GD030", "13/08/2026", "SP002", "Laptop HP Envy 13", "Nhập Kho", 12, 12, "Admin", "Bổ sung tồn kho"],
    ["GD031", "14/08/2026", "SP060", "Cáp HDMI 2.1 8K Baseus 2m", "Nhập Kho", 150, 150, "Admin", "Nhập số lượng lớn"],
    ["GD032", "15/08/2026", "SP061", "Balo chống sốc Targus 15.6 inch", "Nhập Kho", 45, 45, "Admin", "Bổ sung tồn kho"],
    ["GD033", "16/08/2026", "SP065", "Adapter Anker USB-C to HDMI", "Nhập Kho", 80, 80, "Admin", "Nhập số lượng lớn"],
    ["GD034", "17/08/2026", "SP040", "Tai nghe gaming HyperX Cloud II", "Nhập Kho", 40, 40, "Admin", "Bổ sung tồn kho"],
    ["GD035", "18/08/2026", "SP043", "Loa Edifier S880DB", "Nhập Kho", 15, 15, "Admin", "Đặt hàng theo yêu cầu"]
]
# -------------------------------------------------------------
# 8. TẠO CÁC FILE EXCEL ĐỊNH DẠNG CHUYÊN NGHIỆP
# -------------------------------------------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Khởi tạo file Excel riêng biệt cho Bài 7
wb_bt7 = openpyxl.Workbook()
wb_bt7.remove(wb_bt7.active)

sheets_spec = [
    ("DoanhThu_BT1", headers_bt1, branches, "1B365D"),
    ("DonHang_BT2", headers_bt2, orders_flat_for_excel, "005A9C"),
    ("BangLuong_BT3", headers_bt3, payroll, "1B365D"),
    ("DonNghiPhep_BT4", headers_bt4, leaves, "005A9C"),
    ("RawData_BT5", headers_bt5, raw_rows, "4A5568"),
    ("DonHang_BT6", headers_bt6, rfm_rows, "1B365D"),
    ("DanhMuc_BT7", headers_bt7_categories, categories_rows, "7C3AED"),
    ("SanPham_BT7", headers_bt7_products, products_rows, "2563EB"),
    ("KhachHang_BT7", headers_bt7_customers, customers_rows, "059669"),
    ("DonHang_BT7", headers_bt7_orders, orders_master, "DC2626"),
    ("ChiTietDonHang_BT7", headers_bt7_order_details, order_details_rows, "EA580C"),
    ("LichSuTonKho_BT7", headers_bt7_inventory, inventory_rows, "4A5568")
]

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def create_and_style_sheet(workbook, title, headers, rows_data, color_hex):
    ws = workbook.create_sheet(title=title)
    ws.views.sheetView[0].showGridLines = True

    # Title Banner (Merge rộng bằng số lượng cột)
    banner_end_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{banner_end_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"DỮ LIỆU THỰC HÀNH: {title.upper()}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # Headers tại dòng 3
    header_row_idx = 3
    ws.row_dimensions[header_row_idx].height = 24
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ghi dữ liệu từ dòng 4
    current_row = 4
    is_alt_row = False
    last_group_id = None
    for row_idx, r in enumerate(rows_data):
        ws.row_dimensions[current_row].height = 20
        
        # Nếu cột đầu tiên (Mã NV/Mã đơn) có giá trị và khác với nhóm trước đó, đảo màu xen kẽ
        if r[0] is not None and str(r[0]).strip() != "":
            if str(r[0]).strip() != last_group_id:
                is_alt_row = not is_alt_row
                last_group_id = str(r[0]).strip()
                
        bg_fill = PatternFill(start_color="F8FAFC" if is_alt_row else "FFFFFF", fill_type="solid")
        for col_idx, val in enumerate(r, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.fill = bg_fill

            # Định dạng hiển thị số tiền/số thường
            if isinstance(val, (int, float)) and val > 1000:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

    # Merge cells đặc thù cho DonHang_BT2
    if title == "DonHang_BT2":
        start_r = 4
        for order in orders_structured:
            num_items = len(order["items"])
            end_r = start_r + num_items - 1
            if num_items > 1:
                # Xử lý viền cho TẤT CẢ các cột trong dải dòng gộp (12 cột)
                for r in range(start_r, end_r + 1):
                    for col_idx in range(1, 13):  # 12 cột từ A đến L
                        cell = ws.cell(row=r, column=col_idx)
                        # Xác định viền
                        left_border = Side(style='thin', color='D9D9D9')
                        right_border = Side(style='thin', color='D9D9D9')
                        top_border = Side(style='thin', color='D9D9D9') if r == start_r else Side(style=None)
                        bottom_border = Side(style='thin', color='D9D9D9') if r == end_r else Side(style=None)
                        cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
                
                # Sau đó mới gộp các cột cần gộp
                cols_to_merge = [1, 2, 3, 4, 5, 11, 12]
                for col_idx in cols_to_merge:
                    ws.merge_cells(start_row=start_r, start_column=col_idx, end_row=end_r, end_column=col_idx)
                    merged_cell = ws.cell(row=start_r, column=col_idx)
                    h_align = "center" if col_idx in [1, 2, 4, 11] else "left"
                    merged_cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
            start_r = end_r + 1

    # Auto-fit độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= 3 and cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

# 1. Tạo file Excel mẫu tổng hợp
for title, headers, rows_data, color_hex in sheets_spec:
    create_and_style_sheet(wb, title, headers, rows_data, color_hex)

excel_path = os.path.join(DATA_DIR, "Du_Lieu_Mau_Tong_Hop.xlsx")
wb.save(excel_path)
print(f"Created master Excel workbook successfully at: {excel_path}")

# 2. Tạo file Excel riêng cho bài 7
bt7_sheets = [s for s in sheets_spec if "BT7" in s[0]]
for title, headers, rows_data, color_hex in bt7_sheets:
    create_and_style_sheet(wb_bt7, title, headers, rows_data, color_hex)

excel_path_bt7 = os.path.join(DATA_DIR, "bai_tap_7_quan_ly_ban_hang.xlsx")
wb_bt7.save(excel_path_bt7)
print(f"Created Exercise 7 standalone Excel workbook successfully at: {excel_path_bt7}")
